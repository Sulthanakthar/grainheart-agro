import os
import uuid
import datetime
from decimal import Decimal
from django.conf import settings
from django.utils import timezone
from django.contrib.auth import get_user_model
from celery import shared_task
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from .models import Report
from orders.models import Order
from products.models import Product
from accounts.models import Dealer, Customer, Territory
from crm.models import Lead, CustomerInteraction
from payments.models import Payment

User = get_user_model()

def style_header(ws, columns):
    header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='1F4E78', end_color='1F4E78', fill_type='solid')  # Navy Blue
    thin_border = Border(
        left=Side(style='thin', color='CCCCCC'),
        right=Side(style='thin', color='CCCCCC'),
        top=Side(style='thin', color='CCCCCC'),
        bottom=Side(style='thin', color='CCCCCC')
    )
    ws.append(columns)
    for col_num in range(1, len(columns) + 1):
        cell = ws.cell(row=1, column=col_num)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = thin_border
    ws.row_dimensions[1].height = 25

def auto_fit_columns(ws):
    for col in ws.columns:
        max_len = 0
        col_letter = openpyxl.utils.get_column_letter(col[0].column)
        for cell in col:
            val_str = str(cell.value or '')
            if len(val_str) > max_len:
                max_len = len(val_str)
        ws.column_dimensions[col_letter].width = max(max_len + 3, 12)

@shared_task
def generate_report_task(user_id, report_type, start_date_str=None, end_date_str=None):
    # Setup directories
    reports_dir = os.path.join(settings.MEDIA_ROOT, 'reports')
    os.makedirs(reports_dir, exist_ok=True)
    
    # Parse dates
    start_date = None
    end_date = None
    if start_date_str:
        start_date = timezone.make_aware(datetime.datetime.strptime(start_date_str, "%Y-%m-%d"))
    if end_date_str:
        end_date = timezone.make_aware(datetime.datetime.strptime(end_date_str, "%Y-%m-%d") + datetime.timedelta(days=1))

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"{report_type.capitalize()} Report"
    
    user = User.objects.filter(id=user_id).first()
    
    if report_type == 'sales':
        columns = ["Order Number", "Customer", "Order Type", "Subtotal (₹)", "Discount (₹)", "Tax Amount (₹)", "Total Amount (₹)", "Order Status", "Payment Status", "Delivery Address", "Created At"]
        style_header(ws, columns)
        
        queryset = Order.objects.all().select_related('customer').order_by('-created_at')
        if start_date:
            queryset = queryset.filter(created_at__gte=start_date)
        if end_date:
            queryset = queryset.filter(created_at__lt=end_date)
            
        for order in queryset:
            created_at_str = order.created_at.strftime("%Y-%m-%d %H:%M:%S") if order.created_at else ""
            ws.append([
                order.order_number,
                order.customer.username if order.customer else "",
                order.get_order_type_display(),
                float(order.subtotal),
                float(order.discount),
                float(order.tax_amount),
                float(order.total_amount),
                order.get_order_status_display(),
                order.get_payment_status_display(),
                order.delivery_address,
                created_at_str
            ])
            
    elif report_type == 'inventory':
        columns = ["Product Name", "SKU", "Category", "Quality Grade", "Price (₹)", "Available Stock", "Reserved Stock", "Reorder Level", "Weight (kg)", "Is Active"]
        style_header(ws, columns)
        
        queryset = Product.objects.all().select_related('category', 'quality_grade', 'inventory').order_by('sku')
        
        for prod in queryset:
            avail = prod.inventory.available_stock if hasattr(prod, 'inventory') else prod.stock
            res = prod.inventory.reserved_stock if hasattr(prod, 'inventory') else 0
            reorder = prod.inventory.reorder_level if hasattr(prod, 'inventory') else 10
            
            ws.append([
                prod.name,
                prod.sku,
                prod.category.name if prod.category else "",
                prod.quality_grade.name if prod.quality_grade else "",
                float(prod.price),
                avail,
                res,
                reorder,
                float(prod.weight),
                "Yes" if prod.is_active else "No"
            ])
            
    elif report_type == 'dealer':
        columns = ["Dealer Code", "Business Name", "Owner Name", "Phone", "Email", "GST Number", "PAN Number", "Territory", "Commission Rate (%)", "Status", "Approval Date"]
        style_header(ws, columns)
        
        queryset = Dealer.objects.all().select_related('territory').order_by('dealer_code')
        
        for dealer in queryset:
            app_date = dealer.approval_date.strftime("%Y-%m-%d %H:%M:%S") if dealer.approval_date else ""
            ws.append([
                dealer.dealer_code,
                dealer.business_name,
                dealer.owner_name,
                dealer.phone,
                dealer.email,
                dealer.gst_number,
                dealer.pan_number,
                dealer.territory.territory_name if dealer.territory else "",
                float(dealer.commission_rate),
                dealer.get_status_display(),
                app_date
            ])
            
    elif report_type == 'crm':
        columns = ["Lead Number", "Customer Name", "Email", "Phone", "Enquiry Type", "Lead Status", "Priority", "Expected Revenue (₹)", "Conversion Probability (%)", "Assigned To", "Created At"]
        style_header(ws, columns)
        
        queryset = Lead.objects.all().select_related('enquiry', 'assigned_to').order_by('-created_at')
        if start_date:
            queryset = queryset.filter(created_at__gte=start_date)
        if end_date:
            queryset = queryset.filter(created_at__lt=end_date)
            
        for lead in queryset:
            created_at_str = lead.created_at.strftime("%Y-%m-%d %H:%M:%S") if lead.created_at else ""
            ws.append([
                lead.lead_number,
                lead.enquiry.customer_name if lead.enquiry else "",
                lead.enquiry.email if lead.enquiry else "",
                lead.enquiry.phone if lead.enquiry else "",
                lead.enquiry.get_enquiry_type_display() if lead.enquiry else "",
                lead.get_lead_status_display(),
                lead.get_priority_display(),
                float(lead.expected_revenue),
                float(lead.conversion_probability),
                lead.assigned_to.username if lead.assigned_to else "",
                created_at_str
            ])
            
    elif report_type == 'customer':
        columns = ["Username", "Email", "Phone", "City", "State", "Country", "Created At"]
        style_header(ws, columns)
        
        queryset = Customer.objects.all().select_related('user').order_by('-created_at')
        
        for cust in queryset:
            created_at_str = cust.created_at.strftime("%Y-%m-%d %H:%M:%S") if cust.created_at else ""
            ws.append([
                cust.user.username if cust.user else "",
                cust.user.email if cust.user else "",
                cust.phone,
                cust.city,
                cust.state,
                cust.country,
                created_at_str
            ])
            
    elif report_type == 'financial':
        columns = ["Payment ID", "Order Number", "Payment Method", "Transaction Ref", "Amount (₹)", "Payment Status", "Verified By", "Payment Date", "Created At"]
        style_header(ws, columns)
        
        queryset = Payment.objects.all().select_related('order', 'verified_by').order_by('-created_at')
        if start_date:
            queryset = queryset.filter(created_at__gte=start_date)
        if end_date:
            queryset = queryset.filter(created_at__lt=end_date)
            
        for payment in queryset:
            pay_date = payment.payment_date.strftime("%Y-%m-%d %H:%M:%S") if payment.payment_date else ""
            created_at_str = payment.created_at.strftime("%Y-%m-%d %H:%M:%S") if payment.created_at else ""
            ws.append([
                payment.id,
                payment.order.order_number if payment.order else "",
                payment.get_payment_method_display(),
                payment.transaction_reference,
                float(payment.amount),
                payment.get_payment_status_display(),
                payment.verified_by.username if payment.verified_by else "",
                pay_date,
                created_at_str
            ])
            
    elif report_type == 'operational':
        columns = ["Date", "Customer Username", "Lead Number", "Interaction Type", "Description"]
        style_header(ws, columns)
        
        queryset = CustomerInteraction.objects.all().select_related('customer', 'lead').order_by('-interaction_date')
        if start_date:
            queryset = queryset.filter(interaction_date__gte=start_date)
        if end_date:
            queryset = queryset.filter(interaction_date__lt=end_date)
            
        for ix in queryset:
            ix_date = ix.interaction_date.strftime("%Y-%m-%d %H:%M:%S") if ix.interaction_date else ""
            ws.append([
                ix_date,
                ix.customer.username if ix.customer else "",
                ix.lead.lead_number if ix.lead else "",
                ix.get_interaction_type_display(),
                ix.description
            ])
            
    auto_fit_columns(ws)
    
    filename = f"reports/{report_type}_report_{uuid.uuid4().hex[:8]}.xlsx"
    filepath = os.path.join(settings.MEDIA_ROOT, filename)
    wb.save(filepath)
    
    # Create the Report record
    start_label = start_date_str if start_date_str else "All Time"
    end_label = end_date_str if end_date_str else "All Time"
    report_name = f"{report_type.capitalize()} Report ({start_label} to {end_label})"
    
    report_obj = Report.objects.create(
        report_name=report_name,
        report_type=report_type,
        generated_by=user,
        file_path=filename
    )
    
    return report_obj.id
